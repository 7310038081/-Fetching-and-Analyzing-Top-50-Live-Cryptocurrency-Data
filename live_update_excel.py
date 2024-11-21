import requests
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os

# Function to check if the file is a valid Excel file
def is_valid_excel(file_path):
    try:
        # Try to load the workbook
        load_workbook(file_path)
        return True
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return False

# Function to fetch and write data to Excel
def fetch_and_update_excel():
    # URL to fetch data from CoinGecko API
    url = 'https://api.coingecko.com/api/v3/coins/markets'
    params = {
        'vs_currency': 'usd',
        'order': 'market_cap_desc',  # Order by market cap
        'per_page': 50,              # Fetch top 50
        'page': 1                    # Page number
    }

    # Send GET request to the CoinGecko API
    response = requests.get(url, params=params)

    if response.status_code == 200:
        data = response.json()  # Parse JSON data into Python list

        # Create a DataFrame from the fetched data
        df = pd.DataFrame(data)
        
        # Filter and rename necessary columns
        df = df[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h']]
        df.columns = ['Cryptocurrency Name', 'Symbol', 'Current Price (USD)', 
                      'Market Capitalization', '24-Hour Trading Volume (USD)', '24-Hour Price Change (%)']

        # Check if the Excel file exists and is valid
        if not os.path.exists('cryptocurrency_data.xlsx') or not is_valid_excel('cryptocurrency_data.xlsx'):
            print("Creating a new Excel file...")
            # Create a new Excel file if it doesn't exist or is invalid
            writer = pd.ExcelWriter('cryptocurrency_data.xlsx', engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Crypto Data')
            writer.close()  # Close the writer to save the file
            print(f"New Excel file created at {time.strftime('%Y-%m-%d %H:%M:%S')}")
        else:
            # Load the existing Excel file
            try:
                book = load_workbook('cryptocurrency_data.xlsx')
            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                print("Deleting the corrupted file and creating a new one...")
                os.remove('cryptocurrency_data.xlsx')
                # Create a new Excel file after deleting the corrupted one
                writer = pd.ExcelWriter('cryptocurrency_data.xlsx', engine='openpyxl')
                df.to_excel(writer, index=False, sheet_name='Crypto Data')
                writer.close()  # Close the writer to save the file
                print(f"New Excel file created at {time.strftime('%Y-%m-%d %H:%M:%S')}")
                return  # Exit the function as the corrupted file is deleted

            # Create a writer object to append data
            writer = pd.ExcelWriter('cryptocurrency_data.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = {sheet.title: sheet for sheet in book.worksheets}

            # Write data to the 'Crypto Data' sheet
            df.to_excel(writer, index=False, sheet_name='Crypto Data')

            # Save the changes to the Excel file
            writer.close()  # Close the writer to save the file

        print(f"Data updated successfully at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        print(f"Error: Unable to fetch data (status code {response.status_code})")

# Run the script to update the Excel file every 5 minutes
while True:
    fetch_and_update_excel()
    time.sleep(300)  # Wait for 5 minutes (300 seconds)
